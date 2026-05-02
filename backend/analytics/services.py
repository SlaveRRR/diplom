from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from django.db.models import Count, Q
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from analytics.models import AnalyticsEvent
from blog.models import Post
from comics.models import Comic


ALLOWED_INTERVALS = {'day', 'week', 'month'}
EVENT_METRIC_MAP = {
    AnalyticsEvent.EventType.VIEW: 'views',
    AnalyticsEvent.EventType.COMMENT: 'comments',
    AnalyticsEvent.EventType.LIKE: 'likes',
    AnalyticsEvent.EventType.FAVORITE: 'favorites',
    AnalyticsEvent.EventType.PUBLICATION: 'publications',
}
DEFAULT_SUMMARY = {
    'views': 0,
    'reach': 0,
    'comments': 0,
    'likes': 0,
    'favorites': 0,
    'publications': 0,
    'engagement': 0,
    'engagementRate': 0.0,
}


def record_content_event(*, owner, content_kind: str, object_id: int, event_type: str, title_snapshot: str, actor=None):
    if not getattr(owner, 'pk', None):
        return None

    return AnalyticsEvent.objects.create(
        owner=owner,
        actor=actor if getattr(actor, 'pk', None) else None,
        content_kind=content_kind,
        object_id=object_id,
        title_snapshot=title_snapshot[:255],
        event_type=event_type,
    )


def parse_date_range(date_from_raw: str | None, date_to_raw: str | None):
    today = timezone.localdate()
    default_from = today - timedelta(days=29)

    try:
        date_from = datetime.fromisoformat(date_from_raw).date() if date_from_raw else default_from
    except ValueError:
        date_from = default_from

    try:
        date_to = datetime.fromisoformat(date_to_raw).date() if date_to_raw else today
    except ValueError:
        date_to = today

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def get_range_bounds(date_from, date_to):
    current_start = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
    current_end = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    period_days = max((date_to - date_from).days + 1, 1)
    previous_end_date = date_from - timedelta(days=1)
    previous_start_date = previous_end_date - timedelta(days=period_days - 1)
    previous_start = timezone.make_aware(datetime.combine(previous_start_date, datetime.min.time()))
    previous_end = timezone.make_aware(datetime.combine(previous_end_date, datetime.max.time()))

    return current_start, current_end, previous_start, previous_end


def get_available_items(user):
    comics = user.comics.order_by('-updated_at', '-created_at').values('id', 'title', 'status')
    posts = user.posts.order_by('-updated_at', '-created_at').values('id', 'title', 'status')

    return [
        *[
            {
                'id': item['id'],
                'title': item['title'],
                'contentType': AnalyticsEvent.ContentKind.COMIC,
                'status': item['status'],
            }
            for item in comics
        ],
        *[
            {
                'id': item['id'],
                'title': item['title'],
                'contentType': AnalyticsEvent.ContentKind.POST,
                'status': item['status'],
            }
            for item in posts
        ],
    ]


def validate_item_filter(*, user, content_type: str, item_id: int | None):
    if not item_id:
        return None

    if content_type == AnalyticsEvent.ContentKind.COMIC:
        return Comic.objects.filter(author=user, id=item_id).values('id').first()
    if content_type == AnalyticsEvent.ContentKind.POST:
        return Post.objects.filter(author=user, id=item_id).values('id').first()

    return None


def build_event_queryset(*, user, content_type: str, item_id: int | None, start, end):
    queryset = AnalyticsEvent.objects.filter(owner=user, created_at__gte=start, created_at__lte=end)

    if content_type in {AnalyticsEvent.ContentKind.COMIC, AnalyticsEvent.ContentKind.POST}:
        queryset = queryset.filter(content_kind=content_type)

    if item_id:
        queryset = queryset.filter(object_id=item_id)

    return queryset


def summarize_queryset(queryset):
    summary = defaultdict(float)

    aggregated = queryset.aggregate(
        views=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.VIEW)),
        reach=Count('actor', filter=Q(event_type=AnalyticsEvent.EventType.VIEW, actor__isnull=False), distinct=True),
        comments=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.COMMENT)),
        likes=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.LIKE)),
        favorites=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.FAVORITE)),
        publications=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.PUBLICATION)),
    )

    summary.update({key: float(value or 0) for key, value in aggregated.items()})
    summary['engagement'] = summary['comments'] + summary['likes'] + summary['favorites']
    summary['engagementRate'] = round((summary['engagement'] / summary['views']) * 100, 2) if summary['views'] else 0.0
    return summary


def build_metric(current, previous):
    return {
        'value': round(float(current), 2),
        'delta': round(float(current) - float(previous), 2),
    }


def build_summary(*, current_queryset, previous_queryset):
    current = summarize_queryset(current_queryset)
    previous = summarize_queryset(previous_queryset)

    return {
        'views': build_metric(current['views'], previous['views']),
        'reach': build_metric(current['reach'], previous['reach']),
        'comments': build_metric(current['comments'], previous['comments']),
        'likes': build_metric(current['likes'], previous['likes']),
        'favorites': build_metric(current['favorites'], previous['favorites']),
        'publications': build_metric(current['publications'], previous['publications']),
        'engagement': build_metric(current['engagement'], previous['engagement']),
        'engagementRate': build_metric(current['engagementRate'], previous['engagementRate']),
    }


def build_totals_by_content_type(*, user, start, end):
    payload = {}
    for content_type in (AnalyticsEvent.ContentKind.COMIC, AnalyticsEvent.ContentKind.POST):
        summary = summarize_queryset(
            AnalyticsEvent.objects.filter(
                owner=user,
                created_at__gte=start,
                created_at__lte=end,
                content_kind=content_type,
            )
        )
        payload[content_type] = summary
    return payload


def build_timeline(*, queryset, interval: str):
    trunc = TruncDate('created_at')
    if interval == 'week':
        trunc = TruncWeek('created_at')
    elif interval == 'month':
        trunc = TruncMonth('created_at')

    rows = (
        queryset.annotate(bucket=trunc)
        .values('bucket')
        .annotate(
            views=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.VIEW)),
            reach=Count('actor', filter=Q(event_type=AnalyticsEvent.EventType.VIEW, actor__isnull=False), distinct=True),
            comments=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.COMMENT)),
            likes=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.LIKE)),
            favorites=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.FAVORITE)),
            publications=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.PUBLICATION)),
        )
        .order_by('bucket')
    )

    return [
        {
            'period': row['bucket'].strftime('%Y-%m-%d') if row['bucket'] else '',
            'views': row['views'],
            'reach': row['reach'],
            'comments': row['comments'],
            'likes': row['likes'],
            'favorites': row['favorites'],
            'publications': row['publications'],
            'engagement': row['comments'] + row['likes'] + row['favorites'],
        }
        for row in rows
    ]


def build_top_items(queryset):
    rows = (
        queryset.values('content_kind', 'object_id', 'title_snapshot')
        .annotate(
            views=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.VIEW)),
            reach=Count('actor', filter=Q(event_type=AnalyticsEvent.EventType.VIEW, actor__isnull=False), distinct=True),
            comments=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.COMMENT)),
            likes=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.LIKE)),
            favorites=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.FAVORITE)),
            publications=Count('id', filter=Q(event_type=AnalyticsEvent.EventType.PUBLICATION)),
        )
        .order_by('-views', '-comments', '-likes', '-favorites')[:10]
    )

    return [
        {
            'contentType': row['content_kind'],
            'objectId': row['object_id'],
            'title': row['title_snapshot'] or f'{row["content_kind"]} #{row["object_id"]}',
            'views': row['views'],
            'reach': row['reach'],
            'comments': row['comments'],
            'likes': row['likes'],
            'favorites': row['favorites'],
            'publications': row['publications'],
            'engagement': row['comments'] + row['likes'] + row['favorites'],
        }
        for row in rows
    ]


def build_analytics_payload(*, user, content_type: str, item_id: int | None, date_from, date_to, interval: str):
    current_start, current_end, previous_start, previous_end = get_range_bounds(date_from, date_to)
    current_queryset = build_event_queryset(
        user=user,
        content_type=content_type,
        item_id=item_id,
        start=current_start,
        end=current_end,
    )
    previous_queryset = build_event_queryset(
        user=user,
        content_type=content_type,
        item_id=item_id,
        start=previous_start,
        end=previous_end,
    )

    return {
        'filters': {
            'contentType': content_type,
            'itemId': item_id,
            'dateFrom': date_from.isoformat(),
            'dateTo': date_to.isoformat(),
            'interval': interval,
        },
        'summary': build_summary(current_queryset=current_queryset, previous_queryset=previous_queryset),
        'totalsByContentType': build_totals_by_content_type(user=user, start=current_start, end=current_end),
        'timeline': build_timeline(queryset=current_queryset, interval=interval),
        'topItems': build_top_items(current_queryset),
        'availableItems': get_available_items(user),
    }


def build_analytics_workbook(payload):
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = 'Сводка'
    overview_sheet.append(['Метрика', 'Значение', 'Изменение'])
    for cell in overview_sheet[1]:
        cell.font = Font(bold=True)
    metric_labels = {
        'views': 'Просмотры',
        'reach': 'Охват',
        'comments': 'Комментарии',
        'likes': 'Лайки',
        'favorites': 'Избранное',
        'publications': 'Публикации',
        'engagement': 'Вовлечение',
        'engagementRate': 'ER, %',
    }
    for key, value in payload['summary'].items():
        overview_sheet.append([metric_labels.get(key, key), value['value'], value['delta']])

    timeline_sheet = workbook.create_sheet('Динамика')
    timeline_sheet.append(['Период', 'Просмотры', 'Охват', 'Комментарии', 'Лайки', 'Избранное', 'Публикации', 'Вовлечение'])
    for cell in timeline_sheet[1]:
        cell.font = Font(bold=True)
    for point in payload['timeline']:
        timeline_sheet.append([
            point['period'],
            point['views'],
            point['reach'],
            point['comments'],
            point['likes'],
            point['favorites'],
            point['publications'],
            point['engagement'],
        ])

    top_sheet = workbook.create_sheet('Топ материалов')
    top_sheet.append(['Тип контента', 'ID', 'Название', 'Просмотры', 'Охват', 'Комментарии', 'Лайки', 'Избранное', 'Публикации', 'Вовлечение'])
    for cell in top_sheet[1]:
        cell.font = Font(bold=True)
    for item in payload['topItems']:
        top_sheet.append([
            'Комикс' if item['contentType'] == 'comic' else 'Пост',
            item['objectId'],
            item['title'],
            item['views'],
            item['reach'],
            item['comments'],
            item['likes'],
            item['favorites'],
            item['publications'],
            item['engagement'],
        ])

    return workbook


def build_excel_response(payload):
    workbook = build_analytics_workbook(payload)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="analytics-report.xlsx"'
    return response
